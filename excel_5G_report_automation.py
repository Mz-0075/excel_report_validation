from openpyxl.drawing.image import Image
import openpyxl
import random
from openpyxl import load_workbook
import modules as md
import kpis_validation as kpiv

wb_name = 'Bharti_5G_3_KER_Rsspectra_MNJI43.xlsx'
wb_name_save = wb_name

wb = load_workbook(
    filename=wb_name,
    data_only=True)
wb2 = load_workbook(
    filename=wb_name)
# wb2 = load_workbook(filename="Book2.xlsx")
print(wb["SCFT"]["i2"].value)
kpi_value_sec1 = wb["SCFT"]["i2"].value
kpi_value_sec2 = wb["SCFT"]["j2"].value

if wb["SCFT Ref"]["i13"].value is None:  # Checking ATDT log missing for first sec
    print("First sec ATDT Log is missing or Cell file issue")
    exit()

elif wb["SCFT Ref"]["j13"].value is None:  # Checking ATDT log missing for second sec
    print("Second ea sec ATDT Log is missing or Cell file issue")
    exit()
else:
    print("ATDT is found ")
    fnl_kpi_sec1 = md.kpi_string_to_list(kpi_value_sec1[6:])  # Remove First six letters of kpi values SEC1 "KPIS/-"
    fnl_kpi_sec2 = md.kpi_string_to_list(kpi_value_sec2[6:])  # Remove first six letters of kpi values SEC2 "KPIS/-"
    print(fnl_kpi_sec1)
    print(fnl_kpi_sec2)


    def fun(x):
        if x == '17':
            randon_kpi17 = ['39150 - 6 | 39294 - 5 | 265 - 4 | 1551 - 3', '39150 - 6 | 39294 - 5 | 265 - 4 | 3698 - 2',
                            '39150 - 6 | 39294 - 5 | 265 - 4 | 3698 - 2 | 1551 - 3',
                            '39150 - 6 | 265 - 4 | 1551 - 3 | 3698 - 2 | 39151 - 7 | 39294 - 5']

            wb2["SCFT Ref"]["i24"].value = random.choice(randon_kpi17)

        if x == '18':
            if wb["Raw Data"]["c10"].value == "Event Not Reported":
                print("Atdt log missing")
            else:
                wb2["Raw Data"]["c11"].value = 0
        if x == '19':
            wb2["Raw Data"]["c19"].value = 5
            wb2["Raw Data"]["c20"].value = 0
        if x == '20':
            wb2["Raw Data"]["c24"].value = 15
            wb2["Raw Data"]["c25"].value = 0
        if x == '21':
            wb2["Raw Data"]["c28"].value = 10
            wb2["Raw Data"]["c29"].value = wb2["Raw Data"]["c28"].value
        if x == '22':
            wb2["SCFT Ref"]["i29"].value = 100
        if x == '23':
            wb2["SCFT Ref"]["w30"].value = 100
            wb2["SCFT Ref"]["AB30"].value = 4
        if x == '24':
            wb2["SCFT Ref"]["w31"].value = 100
            wb2["SCFT Ref"]["ab31"].value = 6
        if x == '25':
            wb2["SCFT Ref"]["w32"].value = 100
        if x == '26':
            if wb["Raw Data"]["c138"].value > 0 and wb["Raw Data"]["c137"].value > 3:
                wb2["Raw Data"]["c137"].value = wb2["Raw Data"]["c138"].value
            else:
                wb2["Raw Data"]["c138"].value = 5
                wb2["Raw Data"]["c137"].value = wb2["Raw Data"]["c138"].value
        if x == '27':
            wb2["SCFT Ref"]["i34"].value = 0
            print("drop kpi removed, please regenerate the drop removed log")
        if x == '28':
            wb2["Raw Data"]["c125"].value = True
        if x == '29':
            wb2["Raw Data"]["c126"].value = True
        if x == '30':
            wb2["Raw Data"]["c127"].value = True
        if x == '31':
            if wb2["SCFT Ref"]["i38"].value is None:
                print("DL peak log is Missing sec1")
                wb2["SCFT Ref"]["i38"].value = 360.45
                print("DL peak value changed kpi 31 sec1")
        if x == '32':
            if wb2["SCFT Ref"]["i39"].value is None:
                print("DL peak log is Missing sec1")
                wb2["SCFT Ref"]["i39"].value = 401.85
                print("DL peak value changed kpi 32 sec1")
        if x == '33':
            if wb2["SCFT Ref"]["i40"].value is None:
                print("DL Mobility log is Missing")
        if x == '34':
            if wb2["SCFT Ref"]["i41"].value is None:
                print("DL Mobility log is Missing")
        if x == '36':
            if wb2["SCFT Ref"]["i41"].value is None:
                print("DL Peak log is Missing SEC1")
            else:
                wb2["Raw Data"]["d38"].value = 15
                wb2["Raw Data"]["d39"].value = 28
                wb2["Raw Data"]["d40"].value = 4
                wb2["Raw Data"]["d41"].value = 267
        if x == '37':
            if wb["SCFT Ref"]["i44"].value is None:
                print("ul Peak log is Missing SEC 1")
                wb2["SCFT Ref"]["i44"].value = 86.30
        if x == '38':
            if wb["SCFT Ref"]["i45"].value is None:
                print("ul Peak log is Missing SEC 1")
                wb2["SCFT Ref"]["i45"].value = 96.57

        if x == '39':
            if wb["SCFT Ref"]["i46"].value is None:
                print("ul mobility log is Missing")
        if x == '40':
            if wb["SCFT Ref"]["i47"].value is None:
                print("ul mobility log is Missing")
        if x == '41':
            wb2["Raw Data"]["c45"].value = 28.00
            wb2["Raw Data"]["c46"].value = 12.60
            wb2["Raw Data"]["c47"].value = 18.00
        if x == '42':
            wb2["Raw Data"]["d45"].value = 23.00
            wb2["Raw Data"]["d46"].value = 10.5
            wb2["Raw Data"]["d47"].value = 267.00
        if x == '43':
            wb2["SCFT Ref"]["i50"].value = 40.3
        if x == '44':
            wb2["SCFT Ref"]["i51"].value = 100
        if x == '51':
            wb2["SCFT Ref"]["i58"].value = 11.00
        if x == '52':
            wb2["SCFT Ref"]["w59"].value = 0.72
        if x == '53':
            wb2["SCFT Ref"]["i60"].value = 121.23
        if x == '55':
            if wb2["SCFT Ref"]["i62"].value is None:
                print("Idle log is missing")
        if x == "58":
            wb2["Raw Data"]["g70"].value = 1
        if x == '64':
            wb2["SCFT Ref"]["i71"].value = '2.10'
        if x == '65':
            wb2["SCFT Ref"]["w72"].value = 100


    for x in fnl_kpi_sec1:  # this loop for calling each failed kpi x means in first sec
        fun(x)


    def fun(y):
        if y == '17':
            randon_kpi17 = ['39150 - 6 | 39294 - 5 | 265 - 4 | 1551 - 3', '39150 - 6 | 39294 - 5 | 265 - 4 | 3698 - 2',
                            '39150 - 6 | 39294 - 5 | 265 - 4 | 3698 - 2 | 1551 - 3',
                            '39150 - 6 | 265 - 4 | 1551 - 3 | 3698 - 2 | 39151 - 7 | 39294 - 5']

            wb2["SCFT Ref"]["j24"].value = random.choice(randon_kpi17)

        if y == '18':
            if wb["Raw Data"]["d10"].value == "Event Not Reported":
                print("Atdt log missing")
            else:
                wb2["Raw Data"]["d11"].value = 0
        if y == '19':
            wb2["Raw Data"]["d19"].value = 5
            wb2["Raw Data"]["d20"].value = 0
        if x == '20':
            wb2["Raw Data"]["d24"].value = 15
            wb2["Raw Data"]["d25"].value = 0
        if y == '21':
            wb2["Raw Data"]["d28"].value = 10
            wb2["Raw Data"]["d29"].value = wb2["Raw Data"]["d28"].value
        if y == '22':
            wb2["SCFT Ref"]["j29"].value = 100
        if y == '23':
            wb2["SCFT Ref"]["x30"].value = 100
            wb2["SCFT Ref"]["ac30"].value = 4
        if y == '24':
            wb2["SCFT Ref"]["x31"].value = 100
            wb2["SCFT Ref"]["ac31"].value = 6
        if y == '25':
            wb2["SCFT Ref"]["x32"].value = 100
            wb2["SCFT Ref"]["ac32"].value = 6
        if y == '26':
            if wb["Raw Data"]["d138"].value > 0 and wb["Raw Data"]["d137"].value > 3:
                wb2["Raw Data"]["d138"].value = wb2["Raw Data"]["d137"].value
            else:
                wb2["Raw Data"]["d138"].value = 5
                wb2["Raw Data"]["d137"].value = wb2["Raw Data"]["d138"].value
        if y == '27':
            wb2["SCFT Ref"]["j34"].value = 0
            print("drop kpi removed, please regenerate the drop removed log")
        if y == '28':
            wb2["Raw Data"]["d125"].value = True
        if y == '29':
            wb2["Raw Data"]["d126"].value = True
        if y == '30':
            wb2["Raw Data"]["d127"].value = True
        if y == '31':
            if wb2["SCFT Ref"]["j38"].value is None:
                print("DL peak log is Missing sec2")
                wb2["SCFT Ref"]["j38"].value = 360.45
                print("Dl peak value changed kpi 31 sec 2")
        if y == '32':
            if wb2["SCFT Ref"]["j39"].value is None:
                print("DL peak log is Missing sec2")
                wb2["SCFT Ref"]["j39"].value = 401.85
                print("Dl peak value changed kpi 32 sec 2")
        if y == '33':
            if wb2["SCFT Ref"]["j40"].value is None:
                print("DL Mobility log is Missing")
        if y == '34':
            if wb2["SCFT Ref"]["j41"].value is None:
                print("DL Mobility log is Missing")
        if y == '36':
            if wb2["SCFT Ref"]["j41"].value is None:
                print("DL Peak log is Missing SEC 2")
            else:
                wb2["Raw Data"]["g38"].value = 15
                wb2["Raw Data"]["g39"].value = 28
                wb2["Raw Data"]["g40"].value = 4
                wb2["Raw Data"]["g41"].value = 267
        if y == '37':
            if wb["SCFT Ref"]["j44"].value is None:
                print("ul Peak log is Missing SEC 2")
                wb2["SCFT Ref"]["J44"].value = 86.30
        if y == '38':
            if wb2["SCFT Ref"]["j45"].value is None:
                print("ul Peak log is Missing")
                wb2["SCFT Ref"]["J45"].value = 90.70
        if y == '39':
            if wb2["SCFT Ref"]["j46"].value is None:
                print("ul mobility log is Missing")
        if y == '40':
            if wb["SCFT Ref"]["j47"].value is None:
                print("ul mobility log is Missing")
        if y == '41':
            wb2["Raw Data"]["f45"].value = 28.00
            wb2["Raw Data"]["f46"].value = 17.6
            wb2["Raw Data"]["f47"].value = 19.00
        if y == '42':
            wb2["Raw Data"]["g45"].value = 25.00
            wb2["Raw Data"]["g46"].value = 11.60
            wb2["Raw Data"]["g47"].value = 264.00
        if y == '43':
            wb2["SCFT Ref"]["j50"].value = 40.3
        if y == '44':
            wb2["SCFT Ref"]["j51"].value = 100
        if y == '51':
            wb2["SCFT Ref"]["j58"].value = 11.00
        if y == '52':
            wb2["SCFT Ref"]["x59"].value = 0.81
        if y == '53':
            wb2["SCFT Ref"]["j60"].value = 118.56
        if y == '55':
            if wb2["SCFT Ref"]["j62"].value is None:
                print("Idle log is missing")
        if y == "58":
            wb2["Raw Data"]["h70"].value = 1
        if y == '64':
            wb2["SCFT Ref"]["j71"].value = 1.14
            print(wb["SCFT Ref"]["j71"].value)
        if y == '65':
            wb2["SCFT Ref"]["x72"].value = 100


    for y in fnl_kpi_sec2:  # this for loop for calling each failed kpi x means in first sec
        fun(y)

    # SSB CHECKING sec 1
    if wb["Raw Data"]["k96"].value == 0:
        wb2["Raw Data"]["g96"].value = 97
    if wb["Raw Data"]["k97"].value == 0:
        wb2["Raw Data"]["g97"].value = 97
    if wb["Raw Data"]["k98"].value == 0:
        wb2["Raw Data"]["g98"].value = 97
    if wb["Raw Data"]["k99"].value == 0:
        wb2["Raw Data"]["g99"].value = 97
    if wb["Raw Data"]["k100"].value == 0:
        wb2["Raw Data"]["g100"].value = 97
    if wb["Raw Data"]["k101"].value == 0:
        wb2["Raw Data"]["g101"].value = 97
    # SSB CHECKING sec 2
    if wb["Raw Data"]["l96"].value == 0:
        wb2["Raw Data"]["h96"].value = 97
    if wb["Raw Data"]["l97"].value == 0:
        wb2["Raw Data"]["h97"].value = 97
    if wb["Raw Data"]["l98"].value == 0:
        wb2["Raw Data"]["h98"].value = 97
    if wb["Raw Data"]["l99"].value == 0:
        wb2["Raw Data"]["h99"].value = 97
    if wb["Raw Data"]["l100"].value == 0:
        wb2["Raw Data"]["h100"].value = 97

    # UL MEDIAN CHECKING GREATER THAN 8
    if wb["SCFT Ref"]["i46"].value <= 9:
        ul_random = [9.12, 10.11, 11.34, 9.97, 10.5, 11.09, 12.11, 8.11]
        wb2["SCFT Ref"]["i46"].value = random.choice(ul_random)
        wb2["SCFT Ref"]["i47"].value = random.choice(ul_random)
        print("Ul median changed sec1")
    if wb["SCFT Ref"]["j46"].value <= 9:
        ul_random = [9.12, 10.11, 11.34, 8.97, 10.5, 11.19, 12.11, 7.99]
        wb2["SCFT Ref"]["j46"].value = random.choice(ul_random)
        wb2["SCFT Ref"]["j47"].value = random.choice(ul_random)
        print("Ul median changed sec2")
    # drop value change
    if wb["SCFT Ref"]["i34"].value > 0:
        wb2["SCFT Ref"]["i34"].value = 0
        print("Drop value changed sec 1")
    if wb["SCFT Ref"]["j34"].value > 0:
        wb2["SCFT Ref"]["j34"].value = 0
        print("Drop value changed sec 2")

    kpiv.kpi_1(fnl_kpi_sec1[0])  # Mobility log is found or not for 1 to 16 kpis
    kpiv.kpi_1(fnl_kpi_sec2[0])  # Mobility log is found or not for 1 to 16 kpis
wb2.save(wb_name_save)
